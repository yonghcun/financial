from django.shortcuts import render, get_object_or_404, redirect
from .models import MonthlyPerformance
from django.http import HttpResponse
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from io import BytesIO
import os
import pandas as pd

def export_excel(request):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, 'tracker', 'templates', '재무양식.xlsx')

    wb = load_workbook(template_path)
    ws = wb.active

    records = MonthlyPerformance.objects.filter(year=2024).order_by('month')

    start_col = 3
    for idx, record in enumerate(records):
        col = start_col + idx

        ws.cell(row=2, column=col, value=record.revenue_target)
        ws.cell(row=3, column=col, value=record.revenue_sales)
        ws.cell(row=4, column=col, value=record.revenue_ratio)
        ws.cell(row=5, column=col, value=record.cost_cost)
        ws.cell(row=6, column=col, value=record.cost_goods)
        ws.cell(row=7, column=col, value=record.cost_outsourcing)
        ws.cell(row=9, column=col, value=record.sgna)
        ws.cell(row=10, column=col, value=record.op_target)
        ws.cell(row=11, column=col, value=record.op_operating_profit)
        ws.cell(row=12, column=col, value=record.op_ratio)
        ws.cell(row=13, column=col, value=record.op_margin)

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    response = HttpResponse(
        excel_stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=performance_2024.xlsx'
    return response

def round10(value):
    clean = str(value).replace(',', '')
    return Decimal(clean).quantize(Decimal('1.0000000000'), rounding=ROUND_HALF_UP)

def input_view(request):
    if request.method == 'POST':
        year = int(request.POST['year'])
        month = int(request.POST['month'])

        revenue_target = round10(request.POST['revenue_target'])
        revenue_sales = round10(request.POST['revenue_sales'])
        revenue_ratio = round10(request.POST['revenue_ratio'])

        cost_cost = round10(request.POST['cost_cost'])
        cost_goods = round10(request.POST['cost_goods'])
        cost_outsourcing = round10(request.POST['cost_outsourcing'])
        sgna = round10(request.POST['sgna'])

        op_target = round10(request.POST['op_target'])
        op_operating_profit = round10(request.POST['op_operating_profit'])
        op_ratio = round10(request.POST['op_ratio'])
        op_margin = round10(request.POST['op_margin'])

        MonthlyPerformance.objects.update_or_create(
            year=year,
            month=month,
            defaults={
                'revenue_target': revenue_target,
                'revenue_sales': revenue_sales,
                'revenue_ratio': revenue_ratio,
                'cost_cost': cost_cost,
                'cost_goods': cost_goods,
                'cost_outsourcing': cost_outsourcing,
                'sgna': sgna,
                'op_target': op_target,
                'op_operating_profit': op_operating_profit,
                'op_ratio': op_ratio,
                'op_margin': op_margin,
            }
        )
    return redirect('view')

def update_view(request, id):
    record = get_object_or_404(MonthlyPerformance, pk=id)
    if request.method == 'POST':
        for field in ['year', 'month', 'revenue_target', 'revenue_sales', 'cost_cost', 'cost_goods', 'cost_outsourcing', 'sgna', 'op_target', 'op_operating_profit']:
            setattr(record, field, request.POST[field])

        record.revenue_ratio = round(float(record.revenue_sales) / float(record.revenue_target), 10) if float(record.revenue_target) else 0.0
        record.op_ratio = round(float(record.op_operating_profit) / float(record.op_target), 10) if float(record.op_target) else 0.0
        record.op_margin = round(float(record.op_operating_profit) / float(record.revenue_sales), 10) if float(record.revenue_sales) else 0.0

        record.save()
        return redirect('view')
    return render(request, 'input.html', {'record': record})

def view_view(request):
    records = MonthlyPerformance.objects.all().order_by('-year', '-month')
    return render(request, 'view.html', {'records': records})

def stats_view(request):
    records = MonthlyPerformance.objects.all()
    return render(request, 'stats_view.html', {'records': records})

def dynamic_summary_view(request):
    return render(request, 'dynamic_summary.html')
